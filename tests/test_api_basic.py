from pathlib import Path
from time import monotonic, sleep

from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app.api.v1.reports import _build_filtered_option
from app.config import settings
from app.main import app
from app.services.parse_service import _build_option_for_panel

client = TestClient(app)


def _wait_upload_folder_task(task_id: str, timeout_s: float = 30.0) -> dict:
    deadline = monotonic() + timeout_s
    while monotonic() < deadline:
        resp = client.get(f"/consultant/api/v1/reports/upload-folder/tasks/{task_id}")
        assert resp.status_code == 200, resp.json()
        body = resp.json()
        data = body["data"]
        status = data["status"]
        if status in {"succeeded", "failed"}:
            return data
        sleep(0.05)
    raise AssertionError(f"upload task timeout: {task_id}")


def test_health():
    resp = client.get("/health")
    assert resp.status_code == 200
    body = resp.json()
    assert body["code"] == 0
    assert body["data"]["status"] == "ok"


def test_list_reports_success():
    resp = client.get("/consultant/api/v1/reports")
    assert resp.status_code == 200
    body = resp.json()
    assert body["code"] == 0
    assert isinstance(body["data"]["items"], list)


def test_report_crud(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.ensure_dirs()

    create_payload = {
        "report_key": "crud-demo",
        "name": "CRUD Demo",
        "type": "analytics",
        "status": "active",
        "sections": [],
    }
    create_resp = client.post("/consultant/api/v1/reports", json=create_payload)
    assert create_resp.status_code == 200
    create_body = create_resp.json()
    assert create_body["code"] == 0
    assert create_body["data"]["report_key"] == "crud-demo"
    assert create_body["data"]["saved_at"]

    update_resp = client.patch(
        "/consultant/api/v1/reports/crud-demo",
        json={"name": "CRUD Demo Updated", "status": "active"},
    )
    assert update_resp.status_code == 200
    update_body = update_resp.json()
    assert update_body["code"] == 0

    detail_resp = client.get("/consultant/api/v1/reports/crud-demo")
    assert detail_resp.status_code == 200
    detail_body = detail_resp.json()
    assert detail_body["data"]["payload"]["name"] == "CRUD Demo Updated"

    delete_resp = client.delete("/consultant/api/v1/reports/crud-demo")
    assert delete_resp.status_code == 200
    delete_body = delete_resp.json()
    assert delete_body["data"]["deleted"] is True

    missing_resp = client.get("/consultant/api/v1/reports/crud-demo")
    assert missing_resp.status_code == 404


def test_upload_template_v2_auto_build_report(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.ensure_dirs()

    template_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260403"
        / "chapter1_section3.xlsx"
    )
    assert template_path.exists(), f"template not found: {template_path}"

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "template-v2-demo"},
    )

    assert upload_resp.status_code == 200, upload_resp.json()
    upload_body = upload_resp.json()
    assert upload_body["code"] == 0
    assert upload_body["data"]["report_key"] == "template-v2-demo"
    assert upload_body["data"]["parsed_charts"] > 0

    detail_resp = client.get("/consultant/api/v1/reports/template-v2-demo")
    assert detail_resp.status_code == 200
    detail_body = detail_resp.json()
    payload = detail_body["data"]["payload"]

    assert payload["report_key"] == "template-v2-demo"
    assert payload["chapters"]
    section = payload["chapters"][0]["sections"][0]
    chart = section["content_items"]["charts"][0]

    assert payload["chapters"][0]["title"] == "Portfolio Performance Overview"
    assert section["title"] == "Vintage Origination Trends"
    assert section["section_name"] == "Vintage Origination Trends"

    assert "filters" in chart["meta"]
    assert "filter1" in chart["meta"]["filters"]
    assert "filter2" in chart["meta"]["filters"]
    assert "source_rows" in chart["meta"]
    assert chart["echarts"]["xAxis"]["type"] == "time"
    assert chart["echarts"]["xAxis"]["name"]
    assert chart["echarts"]["yAxis"]["name"]
    assert chart["echarts"].get("series")
    assert "markArea" in chart["echarts"]["series"][0]
    assert "markLine" in chart["echarts"]["series"][0]


def test_section_filter_query_uses_duckdb_rows(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    template_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260403"
        / "chapter1_section3.xlsx"
    )

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "duckdb-filter-demo"},
    )
    assert upload_resp.status_code == 200, upload_resp.json()

    all_resp = client.get(
        "/consultant/api/v1/reports/duckdb-filter-demo/chapters/chapter_1/sections/section_1",
        params={"filter1": "All", "filter2": "All"},
    )
    default_resp = client.get(
        "/consultant/api/v1/reports/duckdb-filter-demo/chapters/chapter_1/sections/section_1",
    )
    grade_resp = client.get(
        "/consultant/api/v1/reports/duckdb-filter-demo/chapters/chapter_1/sections/section_1",
        params={"filter1": "Grade 1", "filter2": "36 Month"},
    )

    assert all_resp.status_code == 200, all_resp.json()
    assert default_resp.status_code == 200, default_resp.json()
    assert grade_resp.status_code == 200, grade_resp.json()

    all_chart = all_resp.json()["data"]["section"]["content_items"]["charts"][0]
    default_chart = default_resp.json()["data"]["section"]["content_items"]["charts"][0]
    grade_chart = grade_resp.json()["data"]["section"]["content_items"]["charts"][0]
    default_section = default_resp.json()["data"]["section"]

    assert all_chart["meta"]["filtered_rows_count"] > 0
    assert default_chart["meta"]["filtered_rows_count"] == all_chart["meta"]["filtered_rows_count"]
    assert default_chart["meta"]["selected_filters"] == {"filter1": "ALL", "filter2": "ALL"}
    assert default_section["meta"]["selected_filters"] == {"filter1": "ALL", "filter2": "ALL"}
    assert grade_chart["meta"]["filtered_rows_count"] > 0
    assert grade_chart["meta"]["selected_filters"] == {"filter1": "Grade 1", "filter2": "36 Month"}


def test_upload_table_template_auto_build_report(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    template_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260403"
        / "chapter1_section1.xlsx"
    )
    assert template_path.exists(), f"template not found: {template_path}"

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "table-template-demo"},
    )

    assert upload_resp.status_code == 200, upload_resp.json()

    detail_resp = client.get("/consultant/api/v1/reports/table-template-demo")
    assert detail_resp.status_code == 200
    payload = detail_resp.json()["data"]["payload"]

    chart = payload["chapters"][0]["sections"][0]["content_items"]["charts"][0]
    section = payload["chapters"][0]["sections"][0]
    assert section["section_name"] == section["title"]
    assert chart["chart_type"] == "table"
    assert chart["table_data"]["rows"]
    assert isinstance(chart["table_data"].get("presentation"), dict)
    assert chart["table_data"]["presentation"].get("header_groups")
    assert chart["table_data"]["presentation"].get("cell_styles")
    assert chart["meta"]["filters"]["filter1"]
    assert chart["meta"]["source_rows"]

    section_resp = client.get(
        "/consultant/api/v1/reports/table-template-demo/chapters/chapter_1/sections/section_1",
        params={"filter1": "All", "filter2": "All"},
    )
    assert section_resp.status_code == 200, section_resp.json()
    filtered_chart = section_resp.json()["data"]["section"]["content_items"]["charts"][0]
    assert isinstance(filtered_chart.get("table_data"), dict)
    assert isinstance(filtered_chart["table_data"].get("presentation"), dict)
    assert filtered_chart["table_data"]["presentation"].get("cell_styles")


def test_upload_folder_mixed_templates(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    root = Path(__file__).resolve().parents[2] / "data" / "report20260403"
    table_path = root / "chapter1_section1.xlsx"
    chart_path = root / "chapter1_section3.xlsx"
    assert table_path.exists(), f"template not found: {table_path}"
    assert chart_path.exists(), f"template not found: {chart_path}"

    files = [
        (
            "files",
            (
                table_path.name,
                table_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
        (
            "files",
            (
                chart_path.name,
                chart_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
    ]

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-folder",
        files=files,
        data={"report_key": "folder-mixed-demo", "report_name": "Folder Mixed Demo", "mode": "replace"},
    )
    assert upload_resp.status_code == 200, upload_resp.json()
    body = upload_resp.json()
    assert body["data"]["report_key"] == "folder-mixed-demo"
    assert body["data"]["status"] == "queued"
    task_id = body["data"]["task_id"]

    task_data = _wait_upload_folder_task(task_id)
    assert task_data["status"] == "succeeded"
    assert task_data["succeeded_files"] == 2
    assert task_data["failed_files"] == 0
    assert task_data["result"]["report_key"] == "folder-mixed-demo"

    detail_resp = client.get("/consultant/api/v1/reports/folder-mixed-demo")
    assert detail_resp.status_code == 200
    payload = detail_resp.json()["data"]["payload"]
    chapters = payload["chapters"]
    assert chapters

    all_sections = []
    for chapter in chapters:
        all_sections.extend(chapter.get("sections", []))

    section_keys = {sec.get("section_key") for sec in all_sections if isinstance(sec, dict)}
    assert "section_1" in section_keys
    assert "section_3" in section_keys

    section_titles = {
        sec.get("section_key"): sec.get("title")
        for sec in all_sections
        if isinstance(sec, dict)
    }
    section_names = {
        sec.get("section_key"): sec.get("section_name")
        for sec in all_sections
        if isinstance(sec, dict)
    }
    assert section_names.get("section_1") == section_titles.get("section_1")
    assert section_names.get("section_3") == section_titles.get("section_3")

    upload_subdir = settings.upload_dir / "folder-mixed-demo"
    assert upload_subdir.exists()
    assert (upload_subdir / table_path.name).exists()
    assert (upload_subdir / chart_path.name).exists()


def test_section_filter_does_not_mix_rows_across_chapters(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    root = Path(__file__).resolve().parents[2] / "data" / "report20260403"
    chapter1_path = root / "chapter1_section4.xlsx"
    chapter2_path = root / "chapter2_section4.xlsx"
    assert chapter1_path.exists(), f"template not found: {chapter1_path}"
    assert chapter2_path.exists(), f"template not found: {chapter2_path}"

    files = [
        (
            "files",
            (
                chapter1_path.name,
                chapter1_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
        (
            "files",
            (
                chapter2_path.name,
                chapter2_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
    ]

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-folder",
        files=files,
        data={
            "report_key": "chapter-mix-guard-demo",
            "report_name": "Chapter Mix Guard Demo",
            "mode": "replace",
        },
    )
    assert upload_resp.status_code == 200, upload_resp.json()
    task_id = upload_resp.json()["data"]["task_id"]
    task_data = _wait_upload_folder_task(task_id)
    assert task_data["status"] == "succeeded"

    section_resp = client.get(
        "/consultant/api/v1/reports/chapter-mix-guard-demo/chapters/chapter_1/sections/section_4",
        params={"filter1": "All", "filter2": "All"},
    )
    section_resp_ch2 = client.get(
        "/consultant/api/v1/reports/chapter-mix-guard-demo/chapters/chapter_2/sections/section_4",
        params={"filter1": "All", "filter2": "All"},
    )
    assert section_resp.status_code == 200, section_resp.json()
    assert section_resp_ch2.status_code == 200, section_resp_ch2.json()

    section = section_resp.json()["data"]["section"]
    section_ch2 = section_resp_ch2.json()["data"]["section"]
    assert section["chapter_key"] == "chapter_1"
    assert section_ch2["chapter_key"] == "chapter_2"

    charts_ch1 = section["content_items"]["charts"]
    charts_ch2 = section_ch2["content_items"]["charts"]
    assert charts_ch1
    assert charts_ch2

    count_ch1 = sum(chart.get("meta", {}).get("filtered_rows_count", 0) for chart in charts_ch1)
    count_ch2 = sum(chart.get("meta", {}).get("filtered_rows_count", 0) for chart in charts_ch2)

    assert count_ch1 > 0
    assert count_ch2 > 0
    assert section != section_ch2


def test_get_section_supports_chapter_route_and_legacy_route_removed(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    root = Path(__file__).resolve().parents[2] / "data" / "report20260403"
    chapter1_path = root / "chapter1_section4.xlsx"
    chapter2_path = root / "chapter2_section4.xlsx"

    files = [
        (
            "files",
            (
                chapter1_path.name,
                chapter1_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
        (
            "files",
            (
                chapter2_path.name,
                chapter2_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        ),
    ]

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-folder",
        files=files,
        data={
            "report_key": "chapter-key-query-demo",
            "report_name": "Chapter Key Query Demo",
            "mode": "replace",
        },
    )
    assert upload_resp.status_code == 200, upload_resp.json()
    task_id = upload_resp.json()["data"]["task_id"]
    task_data = _wait_upload_folder_task(task_id)
    assert task_data["status"] == "succeeded"

    chapter2_resp = client.get(
        "/consultant/api/v1/reports/chapter-key-query-demo/chapters/chapter_2/sections/section_4",
        params={"filter1": "All", "filter2": "All"},
    )
    assert chapter2_resp.status_code == 200, chapter2_resp.json()

    section = chapter2_resp.json()["data"]["section"]
    assert section["chapter_key"] == "chapter_2"
    assert section["content_items"]["charts"]
    assert all(chart.get("meta", {}).get("filtered_rows_count", 0) > 0 for chart in section["content_items"]["charts"])

    legacy_resp = client.get(
        "/consultant/api/v1/reports/chapter-key-query-demo/sections/section_4",
        params={"filter1": "All", "filter2": "All"},
    )
    assert legacy_resp.status_code == 404


def test_build_filtered_option_fallbacks_to_category_for_numeric_x_on_time_axis():
    option = {
        "xAxis": {"type": "time"},
        "series": [],
    }
    rows = [
        {"x": 1, "y": 10.0, "legend": "Curve 1", "type": "line"},
        {"x": 2, "y": 12.0, "legend": "Curve 1", "type": "line"},
        {"x": 3, "y": 14.0, "legend": "Curve 1", "type": "line"},
    ]

    rebuilt = _build_filtered_option(option, rows)

    assert rebuilt["xAxis"]["type"] == "category"
    assert rebuilt["xAxis"]["data"] == [1.0, 2.0, 3.0]
    assert rebuilt["series"]
    assert rebuilt["series"][0]["data"] == [10.0, 12.0, 14.0]


def test_section_filter_keeps_numeric_x_chart_data_without_time_axis_fallback(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    template_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260403"
        / "chapter2_section2.xlsx"
    )
    assert template_path.exists(), f"template not found: {template_path}"

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "numeric-x-fallback-demo"},
    )
    assert upload_resp.status_code == 200, upload_resp.json()

    section_resp = client.get(
        "/consultant/api/v1/reports/numeric-x-fallback-demo/chapters/chapter_1/sections/section_1",
        params={"filter1": "All", "filter2": "All"},
    )
    assert section_resp.status_code == 200, section_resp.json()

    charts = section_resp.json()["data"]["section"]["content_items"]["charts"]
    line_charts = [chart for chart in charts if chart.get("chart_type") != "table"]
    assert line_charts

    chart = line_charts[0]
    assert chart["meta"]["filtered_rows_count"] > 0
    assert chart["meta"]["time_axis_fallback_used"] is False

    x_axis = chart["echarts"]["xAxis"]
    assert x_axis["type"] == "category"

    series = chart["echarts"]["series"]
    assert series
    assert any(len(item.get("data") or []) > 0 for item in series)


def test_parse_service_non_timeseries_numeric_x_builds_non_empty_data():
    rows = [
        {
            "x": 1,
            "y": 2.5,
            "legend": "Curve 1",
            "legend_order": 1,
            "type": "line",
            "shape": "none",
            "line_style": "solid",
            "line_width": 1,
            "point_size": 2,
            "color": "#4472C4",
            "y_format": "%",
        },
        {
            "x": 2,
            "y": 3.5,
            "legend": "Curve 1",
            "legend_order": 1,
            "type": "line",
            "shape": "none",
            "line_style": "solid",
            "line_width": 1,
            "point_size": 2,
            "color": "#4472C4",
            "y_format": "%",
        },
    ]

    option = _build_option_for_panel(rows, "facet")

    assert option["xAxis"]["type"] == "category"
    assert option["xAxis"]["data"] == [1.0, 2.0]
    assert option["series"]
    assert option["series"][0]["data"] == [2.5, 3.5]


def test_upload_excel_rejects_old_template_20260330(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    template_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260330"
        / "chapter1_section3.xlsx"
    )
    assert template_path.exists(), f"template not found: {template_path}"

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                template_path.name,
                template_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "old-template-should-fail"},
    )

    assert upload_resp.status_code == 400
    assert "report20260403" in str(upload_resp.json())


def test_upload_excel_rejects_invalid_chart_data_type(tmp_path):
    runtime_dir = tmp_path / "runtime"
    runtime_dir.mkdir(parents=True, exist_ok=True)
    settings.runtime_dir = runtime_dir
    settings.upload_dir = runtime_dir / "uploads"
    settings.parse_dir = runtime_dir / "parsed"
    settings.reports_dir = runtime_dir / "reports"
    settings.meta_file = runtime_dir / "reports_index.json"
    settings.duckdb_file = runtime_dir / "analytics.duckdb"
    settings.ensure_dirs()

    source_path = (
        Path(__file__).resolve().parents[2]
        / "data"
        / "report20260403"
        / "chapter1_section3.xlsx"
    )
    bad_path = tmp_path / "bad_y.xlsx"
    bad_path.write_bytes(source_path.read_bytes())

    wb = load_workbook(bad_path)
    ws = wb["chart_data"]
    header_cells = {str(cell.value).strip().lower(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value}
    assert "y" in header_cells
    ws.cell(row=2, column=header_cells["y"], value="not-a-float")
    wb.save(bad_path)

    upload_resp = client.post(
        "/consultant/api/v1/reports/upload-excel",
        files={
            "file": (
                bad_path.name,
                bad_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        data={"report_key": "invalid-chart-type"},
    )

    assert upload_resp.status_code == 400
    assert "must be float" in str(upload_resp.json())
